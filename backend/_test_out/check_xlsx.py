from openpyxl import load_workbook
wb = load_workbook(r'''C:\Users\hp\Downloads\atoms traders\v7\backend\_test_out\members_test.xlsx''')
ws = wb.active
assert ws.sheet_view.rightToLeft is True
assert ws.freeze_panes == 'A2'
assert ws.auto_filter.ref
assert ws['B1'].value == 'رقم العضوية'
print('rtl', ws.sheet_view.rightToLeft)
print('freeze', ws.freeze_panes)
print('autofilter', ws.auto_filter.ref)
print('rows', ws.max_row)
print('OK_XLSX')
