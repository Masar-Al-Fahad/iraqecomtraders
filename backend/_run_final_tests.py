"""Final verification tests for permissions, membership numbers, excel, sorting."""
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

BASE = "http://127.0.0.1:8000"
OUT = Path(__file__).resolve().parent / "_test_out"
OUT.mkdir(exist_ok=True)
results = []


def req(method, path, token=None, body=None, raw=False):
    data = None
    headers = {}
    if body is not None:
        data = json.dumps(body).encode("utf-8")
        headers["Content-Type"] = "application/json"
    if token:
        headers["Authorization"] = f"Bearer {token}"
    r = urllib.request.Request(f"{BASE}{path}", data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(r, timeout=30) as resp:
            content = resp.read()
            if raw:
                return resp.status, content, dict(resp.headers)
            return resp.status, json.loads(content.decode("utf-8") or "null"), dict(resp.headers)
    except urllib.error.HTTPError as e:
        content = e.read()
        try:
            parsed = json.loads(content.decode("utf-8") or "null")
        except Exception:
            parsed = content.decode("utf-8", errors="replace")
        if raw:
            return e.code, content, {}
        return e.code, parsed, {}


def login(username, password):
    code, data, _ = req("POST", "/api/v1/auth/login", body={"username": username, "password": password})
    assert code == 200, f"login failed {code} {data}"
    return data["token"], data["user"]


def main():
    admin_token, admin_user = login("admin", "Admin@12345")
    results.append(f"LOGIN_ADMIN: PASS name={admin_user.get('name')} super={admin_user.get('is_super_admin')}")

    uname = "nodelete_pytest"
    # delete existing if any
    code, users, _ = req("GET", "/api/v1/admin/users", token=admin_token)
    assert code == 200, users
    for u in users.get("items", []):
        if u["username"] == uname:
            req("DELETE", f"/api/v1/admin/users/{u['id']}", token=admin_token)

    code, created, _ = req(
        "POST",
        "/api/v1/admin/users",
        token=admin_token,
        body={
            "username": uname,
            "password": "Test@12345",
            "is_active": True,
            "permissions": {
                "view": True,
                "add": True,
                "edit": True,
                "delete": False,
                "export": True,
                "manage_users": False,
            },
        },
    )
    assert code == 200, created
    results.append(f"CREATE_LIMITED_USER: PASS id={created['id']} delete={created['permissions']['delete']}")

    lim_token, lim_user = login(uname, "Test@12345")
    results.append(
        f"LOGIN_LIMITED: PASS delete={lim_user['permissions']['delete']} edit={lim_user['permissions']['edit']}"
    )

    # ensure a member exists
    code, listing, _ = req("GET", "/api/v1/admin/registrations?limit=50&sort=-created_at", token=admin_token)
    assert code == 200, listing
    items = listing.get("items") or []
    if not items:
        code, seed, _ = req(
            "POST",
            "/api/v1/admin/registrations/add-member",
            token=admin_token,
            body={
                "business_name": "Seed Biz",
                "merchant_name": "Seed Merchant",
                "phone": "07701234567",
                "governorate": "بغداد",
                "area": "الكرادة",
                "business_type": "تجارة",
                "notes": "seed",
                "membership_status": "active",
            },
        )
        assert code == 200, seed
        items = [seed]
        results.append(f"SEED: {seed.get('membership_number')}")

    target_id = items[0]["id"]

    # TEST 1: delete without permission
    code, detail, _ = req("DELETE", f"/api/v1/admin/registrations/{target_id}", token=lim_token)
    msg = detail.get("detail") if isinstance(detail, dict) else str(detail)
    ok = code == 403 and msg == "ليس لديك صلاحية لتنفيذ هذا الإجراء."
    results.append(f"TEST1_DELETE_403: status={code} msg={msg} PASS={ok}")

    code, detail, _ = req("GET", "/api/v1/admin/users", token=lim_token)
    msg = detail.get("detail") if isinstance(detail, dict) else str(detail)
    ok = code == 403 and "ليس لديك صلاحية" in str(msg)
    results.append(f"TEST1b_MANAGE_USERS_403: status={code} msg={msg} PASS={ok}")

    # export without export perm
    code, u2, _ = req(
        "POST",
        "/api/v1/admin/users",
        token=admin_token,
        body={
            "username": "noexport_pytest",
            "password": "Test@12345",
            "is_active": True,
            "permissions": {
                "view": True,
                "add": False,
                "edit": False,
                "delete": False,
                "export": False,
                "manage_users": False,
            },
        },
    )
    if code == 200 or (isinstance(u2, dict) and "موجود" in str(u2)):
        # recreate if exists
        if code != 200:
            for u in users.get("items", []):
                if u["username"] == "noexport_pytest":
                    req("DELETE", f"/api/v1/admin/users/{u['id']}", token=admin_token)
            code, u2, _ = req(
                "POST",
                "/api/v1/admin/users",
                token=admin_token,
                body={
                    "username": "noexport_pytest",
                    "password": "Test@12345",
                    "is_active": True,
                    "permissions": {
                        "view": True,
                        "add": False,
                        "edit": False,
                        "delete": False,
                        "export": False,
                        "manage_users": False,
                    },
                },
            )
    noexp_token, _ = login("noexport_pytest", "Test@12345")
    code, detail, _ = req("GET", "/api/v1/admin/registrations/export-xlsx", token=noexp_token)
    msg = detail.get("detail") if isinstance(detail, dict) else str(detail)
    ok = code == 403
    results.append(f"TEST1c_EXPORT_403: status={code} msg={msg} PASS={ok}")

    # Entities API bypass (was unprotected)
    code, detail, _ = req("DELETE", f"/api/v1/entities/registrations/{target_id}", token=lim_token)
    msg = detail.get("detail") if isinstance(detail, dict) else str(detail)
    ok = code == 403
    results.append(f"TEST1d_ENTITIES_DELETE_403: status={code} msg={msg} PASS={ok}")

    # TEST 2 membership sequence
    code, listing, _ = req("GET", "/api/v1/admin/registrations?limit=200&sort=-created_at", token=admin_token)
    items = listing.get("items") or []
    max_n = 0
    delete_candidate = None
    for it in items:
        mn = it.get("membership_number") or ""
        if mn.startswith("MF-"):
            try:
                n = int(mn.split("-", 1)[1])
            except ValueError:
                continue
            max_n = max(max_n, n)
            if delete_candidate is None:
                delete_candidate = it

    if delete_candidate is None:
        code, delete_candidate, _ = req(
            "POST",
            "/api/v1/admin/registrations/add-member",
            token=admin_token,
            body={
                "business_name": "ToDelete",
                "merchant_name": "X",
                "phone": "07701112233",
                "governorate": "بغداد",
                "area": "أ",
                "business_type": "ت",
                "notes": "",
                "membership_status": "active",
            },
        )
        assert code == 200, delete_candidate
        mn = delete_candidate["membership_number"]
        max_n = max(max_n, int(mn.split("-", 1)[1]))

    deleted_mn = delete_candidate["membership_number"]
    code, _, _ = req("DELETE", f"/api/v1/admin/registrations/{delete_candidate['id']}", token=admin_token)
    assert code == 200, "delete failed"

    code, new_m, _ = req(
        "POST",
        "/api/v1/admin/registrations/add-member",
        token=admin_token,
        body={
            "business_name": "New After Delete",
            "merchant_name": "New Merchant",
            "phone": "07709998877",
            "governorate": "بغداد",
            "area": "المنصور",
            "business_type": "تجارة",
            "notes": "mf-test",
            "membership_status": "active",
        },
    )
    assert code == 200, new_m
    expected = f"MF-{max_n + 1:04d}"
    ok = new_m["membership_number"] == expected and new_m["membership_number"] != deleted_mn
    results.append(
        f"TEST2_MF_SEQUENCE: got={new_m['membership_number']} expected={expected} deleted={deleted_mn} PASS={ok}"
    )

    # TEST 3 excel
    code, content, headers = req("GET", "/api/v1/admin/registrations/export-xlsx", token=admin_token, raw=True)
    xlsx_path = OUT / "members_test.xlsx"
    xlsx_path.write_bytes(content)
    sig_ok = content[:2] == b"PK"
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active
    xlsx_ok = (
        code == 200
        and sig_ok
        and ws.sheet_view.rightToLeft is True
        and ws.freeze_panes == "A2"
        and bool(ws.auto_filter.ref)
        and ws["B1"].value == "رقم العضوية"
    )
    results.append(
        f"TEST3_EXCEL: status={code} sig_PK={sig_ok} rtl={ws.sheet_view.rightToLeft} "
        f"freeze={ws.freeze_panes} filter={ws.auto_filter.ref} rows={ws.max_row} PASS={xlsx_ok}"
    )

    # TEST 4 sorting
    fields = [
        "membership_number",
        "business_name",
        "merchant_name",
        "phone",
        "governorate",
        "status",
        "membership_status",
        "created_at",
        "approved_at",
        "last_modified_by",
        "updated_at",
    ]
    sort_ok = True
    for f in fields:
        for sort in (f, f"-{f}"):
            code, data, _ = req("GET", f"/api/v1/admin/registrations?limit=5&sort={sort}", token=admin_token)
            if code != 200 or data.get("items") is None:
                sort_ok = False
                results.append(f"SORT_FAIL {sort} code={code} data={data}")
    results.append(f"TEST4_SORT_ALL: PASS={sort_ok}")

    code, asc, _ = req("GET", "/api/v1/admin/registrations?limit=50&sort=membership_number", token=admin_token)
    nums = []
    for it in asc.get("items") or []:
        mn = it.get("membership_number") or ""
        if mn.startswith("MF-"):
            try:
                nums.append(int(mn.split("-", 1)[1]))
            except ValueError:
                pass
    num_ok = nums == sorted(nums)
    results.append(f"TEST4b_NUMERIC_MF_ASC: {nums} PASS={num_ok}")

    # TEST 5 combo
    code, combo, _ = req(
        "GET",
        "/api/v1/admin/registrations?limit=20&status=approved&sort=-created_at&query=%D8%A8",
        token=admin_token,
    )
    results.append(f"TEST5_COMBO: status={code} total={combo.get('total')} PASS={code == 200}")

    # TEST 6 system
    code_s, stats, _ = req("GET", "/api/v1/admin/registrations/stats", token=admin_token)
    code_u, users2, _ = req("GET", "/api/v1/admin/users", token=admin_token)
    code_c, check, _ = req("GET", "/api/v1/admin/registrations/check-admin", token=admin_token)
    results.append(f"TEST6_STATS: status={code_s} total={stats.get('total')} PASS={code_s == 200}")
    results.append(f"TEST6_USERS: status={code_u} total={users2.get('total')} PASS={code_u == 200}")
    results.append(
        f"TEST6_CHECK_ADMIN: status={code_c} authorized={check.get('authorized')} PASS={check.get('authorized') is True}"
    )

    try:
        with urllib.request.urlopen("http://127.0.0.1:5173/", timeout=5) as r:
            results.append(f"TEST6_FRONTEND: status={r.status} PASS={r.status == 200}")
    except Exception as e:
        results.append(f"TEST6_FRONTEND: FAIL {e}")

    # cleanup test users
    code, users3, _ = req("GET", "/api/v1/admin/users", token=admin_token)
    for u in users3.get("items", []):
        if u["username"] in (uname, "noexport_pytest"):
            req("DELETE", f"/api/v1/admin/users/{u['id']}", token=admin_token)

    text = "\n".join(results)
    (OUT / "results.txt").write_text(text, encoding="utf-8")
    print(text)
    fails = [line for line in results if "PASS=False" in line]
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
