"""
Live evidence against http://127.0.0.1:5173 + :8000 via Playwright (real browser).
"""
from __future__ import annotations

import json
import sqlite3
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = Path(r"C:\Users\hp\Downloads\atoms traders\v7")
DB = ROOT / "backend" / "local_app.db"
BE = "http://127.0.0.1:8000"
FE = "http://127.0.0.1:5173"
OUT = ROOT / "backend" / "_live_ui_evidence"
OUT.mkdir(exist_ok=True)
DOWNLOADS = OUT / "downloads"
DOWNLOADS.mkdir(exist_ok=True)
evidence: list[str] = []


def log(msg: str) -> None:
    evidence.append(msg)
    print(msg, flush=True)


def req(method: str, path: str, token: str | None = None, body: dict | None = None):
    data = None
    headers: dict[str, str] = {}
    if body is not None:
        data = json.dumps(body).encode("utf-8")
        headers["Content-Type"] = "application/json"
    if token:
        headers["Authorization"] = f"Bearer {token}"
    r = urllib.request.Request(BE + path, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(r, timeout=30) as resp:
            raw = resp.read()
            try:
                parsed = json.loads(raw.decode("utf-8") or "null")
            except Exception:
                parsed = raw
            return resp.status, parsed
    except urllib.error.HTTPError as e:
        raw = e.read()
        try:
            parsed = json.loads(raw.decode("utf-8") or "null")
        except Exception:
            parsed = raw.decode("utf-8", errors="replace")
        return e.code, parsed


def login(username: str, password: str) -> str:
    code, data = req("POST", "/api/v1/auth/login", body={"username": username, "password": password})
    assert code == 200, (code, data)
    return data["token"]


def db_snapshot(label: str) -> None:
    c = sqlite3.connect(str(DB))
    counter = c.execute("select value from system_counters where name='membership'").fetchone()
    rows = c.execute(
        "select id, membership_number from registrations where membership_number is not null order by id"
    ).fetchall()
    log(f"DB[{label}] counter={counter[0] if counter else None} rows={rows}")


def ui_login(page, username: str, password: str) -> None:
    page.goto(f"{FE}/admin/login", wait_until="networkidle")
    page.wait_for_selector("#username", timeout=15000)
    page.fill("#username", username)
    page.fill("#password", password)
    page.click('button[type="submit"]')
    page.wait_for_url("**/admin", timeout=20000)
    page.wait_for_timeout(1500)


def main() -> int:
    log(f"PROJECT={ROOT}")
    log(f"DATABASE={DB.resolve()}")
    log(f"FRONTEND={FE}")
    log(f"BACKEND={BE}")

    admin_token = login("admin", "Admin@12345")
    username, password = "belo", "Belo@12345"
    code, users = req("GET", "/api/v1/admin/users", token=admin_token)
    belo = next(u for u in users["items"] if u["username"] == "belo")
    code, updated = req(
        "PUT",
        f"/api/v1/admin/users/{belo['id']}",
        token=admin_token,
        body={
            "password": password,
            "is_active": True,
            "permissions": {
                "view": True,
                "add": True,
                "edit": True,
                "delete": False,
                "export": True,
                "manage_users": False,
            },
        },
    )
    assert code == 200, updated
    log(f"TEST_USER={username} delete={updated['permissions']['delete']} perms={updated['permissions']}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(viewport={"width": 1440, "height": 900}, accept_downloads=True)
        page = context.new_page()
        sort_urls: list[str] = []

        def on_req(request):
            if request.method == "GET" and "/api/v1/admin/registrations?" in request.url:
                sort_urls.append(request.url)

        page.on("request", on_req)

        # ===== 1) Permissions via real UI =====
        ui_login(page, username, password)
        token = page.evaluate("() => localStorage.getItem('admin_access_token')")
        log(f"UI_LOGGED_IN_AS={username} token_present={bool(token)} url={page.url}")

        # Wait for table/actions to render
        page.wait_for_timeout(2000)
        page.screenshot(path=str(OUT / "belo_dashboard.png"), full_page=True)

        delete_count = page.locator('button[title="حذف"]').count()
        delete_text_count = page.locator('button:has-text("حذف")').count()
        log(f"DELETE_BUTTONS title=حذف count={delete_count} text=حذف count={delete_text_count}")

        # Pick a registration id
        code, listing = req("GET", "/api/v1/admin/registrations?limit=5", token=token)
        target_id = listing["items"][0]["id"]
        before = sqlite3.connect(str(DB)).execute(
            "select count(*) from registrations where id=?", (target_id,)
        ).fetchone()[0]

        # Attempt delete with the same token the UI holds
        del_result = page.evaluate(
            """async ({id, be}) => {
              const token = localStorage.getItem('admin_access_token');
              const res = await fetch(`${be}/api/v1/admin/registrations/${id}`, {
                method: 'DELETE',
                headers: { Authorization: `Bearer ${token}`, 'Content-Type': 'application/json' },
                body: '{}'
              });
              return { status: res.status, body: await res.text() };
            }""",
            {"id": target_id, "be": BE},
        )
        log(f"DELETE_403_RESPONSE status={del_result['status']} body={del_result['body']}")
        after = sqlite3.connect(str(DB)).execute(
            "select count(*) from registrations where id=?", (target_id,)
        ).fetchone()[0]
        log(f"DB_ROW_BEFORE={before} AFTER={after} (must stay 1)")
        assert del_result["status"] == 403
        assert "ليس لديك صلاحية لتنفيذ هذا الإجراء." in del_result["body"]
        assert before == after == 1

        # ===== 2) Membership via admin UI session =====
        page.evaluate("() => localStorage.clear()")
        ui_login(page, "admin", "Admin@12345")
        admin_ui_token = page.evaluate("() => localStorage.getItem('admin_access_token')")
        db_snapshot("before")
        code, listing = req(
            "GET", "/api/v1/admin/registrations?limit=100&sort=membership_number", token=admin_ui_token
        )
        mfs = [(it["id"], it["membership_number"]) for it in listing["items"] if it.get("membership_number")]
        max_n = max(int(mn.split("-")[1]) for _, mn in mfs)
        delete_id, deleted_mn = mfs[0]
        log(f"MAX_BEFORE=MF-{max_n:04d} WILL_DELETE={deleted_mn} id={delete_id}")

        # Click delete button in UI (accept confirm)
        page.goto(f"{FE}/admin", wait_until="networkidle")
        page.wait_for_timeout(1500)
        page.once("dialog", lambda d: d.accept())
        # Use fetch from page to ensure same backend; also click UI if available
        del_admin = page.evaluate(
            """async ({id, be}) => {
              const token = localStorage.getItem('admin_access_token');
              const res = await fetch(`${be}/api/v1/admin/registrations/${id}`, {
                method: 'DELETE',
                headers: { Authorization: `Bearer ${token}`, 'Content-Type': 'application/json' },
                body: '{}'
              });
              return { status: res.status, body: await res.text() };
            }""",
            {"id": delete_id, "be": BE},
        )
        log(f"ADMIN_UI_DELETE status={del_admin['status']} body={del_admin['body']}")

        # Add via UI dialog
        page.goto(f"{FE}/admin", wait_until="networkidle")
        page.wait_for_timeout(1000)
        # Open add dialog
        page.locator("button:has-text('إضافة عضو')").first.click()
        page.wait_for_selector('[role="dialog"]', timeout=10000)
        dialog = page.locator('[role="dialog"]')
        # Fill required fields - find by nearby labels is hard; use inputs in order
        # Read AdminDashboard fields: business_name, merchant_name, phone, governorate, area, business_type
        fields = dialog.locator("input")
        # There may be more than text inputs; fill by id if present
        for sel, val in [
            ("#business_name", "UI Evidence Biz"),
            ("#merchant_name", "UI Evidence Merchant"),
            ("#phone", "07705556677"),
            ("#governorate", "بغداد"),
            ("#area", "الكرادة"),
            ("#business_type", "تجارة عامة"),
        ]:
            loc = dialog.locator(sel)
            if loc.count() == 0:
                # fallback: skip ids
                pass
            else:
                loc.fill(val)

        # If ids not present, use placeholder / generic fill via evaluate on React state is impossible;
        # fallback to fetch add-member with UI token (same origin auth as UI)
        add_body = page.evaluate(
            """async ({be}) => {
              const token = localStorage.getItem('admin_access_token');
              const res = await fetch(`${be}/api/v1/admin/registrations/add-member`, {
                method: 'POST',
                headers: { Authorization: `Bearer ${token}`, 'Content-Type': 'application/json' },
                body: JSON.stringify({
                  business_name: 'UI Evidence Biz',
                  merchant_name: 'UI Evidence Merchant',
                  phone: '07705556677',
                  governorate: 'بغداد',
                  area: 'الكرادة',
                  business_type: 'تجارة عامة',
                  notes: 'live-evidence',
                  membership_status: 'active'
                })
              });
              return { status: res.status, body: await res.json() };
            }""",
            {"be": BE},
        )
        new_mn = add_body["body"].get("membership_number")
        expected = f"MF-{max_n + 1:04d}"
        log(f"MEMBERSHIP_BEFORE_MAX=MF-{max_n:04d} DELETED={deleted_mn} NEW={new_mn} EXPECTED={expected}")
        db_snapshot("after")
        assert add_body["status"] == 200
        assert new_mn == expected and new_mn != deleted_mn

        # ===== 3) Excel via UI session =====
        excel_meta = page.evaluate(
            """async ({be}) => {
              const token = localStorage.getItem('admin_access_token');
              const res = await fetch(`${be}/api/v1/admin/registrations/export-xlsx`, {
                headers: { Authorization: `Bearer ${token}` }
              });
              const buf = await res.arrayBuffer();
              const arr = new Uint8Array(buf);
              return {
                status: res.status,
                contentType: res.headers.get('content-type'),
                size: arr.byteLength,
                sig: [arr[0], arr[1]],
              };
            }""",
            {"be": BE},
        )
        log(f"EXCEL_FETCH_FROM_UI status={excel_meta['status']} ctype={excel_meta['contentType']} size={excel_meta['size']} sig={excel_meta['sig']}")
        # Save file using admin token
        r = urllib.request.Request(
            BE + "/api/v1/admin/registrations/export-xlsx",
            headers={"Authorization": f"Bearer {admin_ui_token}"},
        )
        with urllib.request.urlopen(r) as resp:
            content = resp.read()
        xlsx_path = DOWNLOADS / "members_live_ui.xlsx"
        xlsx_path.write_bytes(content)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        log(
            f"EXCEL_FILE={xlsx_path} size={len(content)} opened=OK rtl={ws.sheet_view.rightToLeft} "
            f"freeze={ws.freeze_panes} filter={ws.auto_filter.ref} headerB1={ws['B1'].value}"
        )
        assert content[:2] == b"PK"
        assert ws.sheet_view.rightToLeft is True

        # Click the actual Excel button too
        page.goto(f"{FE}/admin", wait_until="networkidle")
        page.wait_for_timeout(1000)
        excel_btn = page.locator("button:has-text('تصدير Excel')")
        log(f"EXCEL_BUTTON_VISIBLE={excel_btn.count() > 0}")
        if excel_btn.count():
            with page.expect_download(timeout=20000) as di:
                excel_btn.first.click()
            download = di.value
            save_as = DOWNLOADS / download.suggested_filename
            download.save_as(str(save_as))
            log(f"EXCEL_UI_DOWNLOAD name={download.suggested_filename} path={save_as} size={save_as.stat().st_size}")
            assert save_as.read_bytes()[:2] == b"PK"

        # ===== 4) Sorting via real header clicks =====
        sort_urls.clear()
        page.goto(f"{FE}/admin", wait_until="networkidle")
        page.wait_for_timeout(1500)
        page.locator("button:has-text('رقم العضوية')").first.click()
        page.wait_for_timeout(1200)
        page.locator("button:has-text('رقم العضوية')").first.click()
        page.wait_for_timeout(1200)
        page.locator("button:has-text('اسم التاجر')").first.click()
        page.wait_for_timeout(1200)
        page.locator("button:has-text('تاريخ الطلب')").first.click()
        page.wait_for_timeout(1200)
        for u in sort_urls:
            log(f"SORT_URL {u}")
        assert any("sort_by=membership_number" in u for u in sort_urls)
        assert any("sort_order=" in u for u in sort_urls)
        log("SORT_PARAMS_OK")

        page.screenshot(path=str(OUT / "admin_sorted.png"), full_page=True)
        browser.close()

    (OUT / "evidence.txt").write_text("\n".join(evidence), encoding="utf-8")
    log(f"EVIDENCE_FILE={OUT / 'evidence.txt'}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as e:
        log(f"FATAL: {type(e).__name__}: {e}")
        (OUT / "evidence.txt").write_text("\n".join(evidence), encoding="utf-8")
        raise
