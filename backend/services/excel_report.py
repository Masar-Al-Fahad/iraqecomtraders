"""Professional MFEC Excel membership report builder (openpyxl)."""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from urllib.parse import urlparse
from urllib.request import urlopen

from services.extra_fields import dynamic_field_defs, extra_field_value

DEFAULT_NAVY = "1F2937"
DEFAULT_GOLD = "C89B3C"
WHITE = "FFFFFF"
ALT_ROW = "F3F4F6"
BORDER = "9CA3AF"

DEFAULT_ORG_NAME = "تجمع تجار التجارة الإلكترونية في العراق"
DEFAULT_ORG_ABBR = "MFEC"
DEFAULT_REPORT_TITLE = "تقرير أعضاء تجمع تجار التجارة الإلكترونية في العراق"
DEFAULT_SPONSOR = "برعاية شركة مسار الفهد للتجارة العامة والنقل العام"
DEFAULT_WEBSITE = "www.masaralfahad.com"
DEFAULT_EMAIL = "management@masaralfahad.com"
DEFAULT_PHONE = "07748077716"

BASE_HEADERS = [
    "#",
    "رقم العضوية",
    "اسم النشاط التجاري",
    "اسم التاجر",
    "رقم الهاتف",
    "المحافظة",
    "المنطقة",
    "نوع النشاط",
    "حالة الطلب",
    "حالة العضوية",
    "تاريخ الطلب",
    "تاريخ الموافقة",
    "آخر تعديل بواسطة",
    "تاريخ آخر تعديل",
    "ملاحظات",
]

STATUS_MAP = {"approved": "مقبول", "rejected": "مرفوض", "pending": "قيد المراجعة"}
MS_MAP = {"active": "فعال", "suspended": "معلق", "expired": "منتهي"}

LOGO_CANDIDATES = [
    Path(__file__).resolve().parent.parent / "assets" / "brand" / "mfec-logo.png",
    Path(__file__).resolve().parent.parent / "static" / "brand" / "mfec-logo.png",
]


def _hex(color: str, fallback: str) -> str:
    c = (color or fallback).lstrip("#").upper()
    return c if len(c) in (6, 8) else fallback


def _logo_path(brand: Optional[Dict[str, Any]] = None) -> Optional[Path]:
    """Resolve a local temp/path for Excel logo embedding.

    Prefer downloading brand-file assets from Supabase Storage public URL.
    Fall back to bundled static logos.
    """
    import tempfile

    from services import supabase_storage as s3store

    brand = brand or {}
    logo = brand.get("report_logo") or brand.get("system_logo") or ""

    # API brand-file → Supabase Storage brand/<name>
    if logo.startswith("/api/v1/public/app-settings/brand-file/"):
        name = logo.rsplit("/", 1)[-1]
        storage_key = s3store.normalize_brand_key(name)
        try:
            # sync download via urllib for openpyxl path convenience
            public = s3store.public_object_url(storage_key)
            with urlopen(public, timeout=20) as resp:
                data = resp.read()
            if data:
                suffix = Path(name).suffix or ".png"
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
                tmp.write(data)
                tmp.close()
                return Path(tmp.name)
        except Exception:
            pass

    if logo.startswith("http://") or logo.startswith("https://"):
        try:
            with urlopen(logo, timeout=20) as resp:
                data = resp.read()
            if data:
                suffix = Path(urlparse(logo).path).suffix or ".png"
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
                tmp.write(data)
                tmp.close()
                return Path(tmp.name)
        except Exception:
            pass

    if logo.startswith("/brand/"):
        p = Path(__file__).resolve().parent.parent.parent / "frontend" / "public" / logo.lstrip("/")
        if p.is_file():
            return p
        p2 = Path(__file__).resolve().parent.parent / "assets" / "brand" / Path(logo).name
        if p2.is_file():
            return p2
    for p in LOGO_CANDIDATES:
        if p.is_file():
            return p
    return None


def _row_values(idx: int, item: Any, dyn_fields: Sequence[Dict[str, Any]], form_settings: Optional[Dict[str, Any]] = None) -> List[Any]:
    values = [
        idx,
        getattr(item, "membership_number", None) or "-",
        item.business_name,
        item.merchant_name,
        item.phone,
        item.governorate,
        item.area,
        getattr(item, "business_type", None) or "-",
        STATUS_MAP.get(item.status, item.status),
        MS_MAP.get(getattr(item, "membership_status", None) or "", getattr(item, "membership_status", None) or "-"),
        str(item.created_at)[:19] if item.created_at else "-",
        (getattr(item, "approved_at", None) or "-")[:19],
        getattr(item, "last_modified_by", None) or "-",
        str(item.updated_at)[:19] if item.updated_at else "-",
        getattr(item, "notes", None) or "-",
    ]
    for f in dyn_fields:
        values.append(extra_field_value(item, str(f.get("id") or ""), form_settings) or "-")
    return values


def build_members_xlsx(
    items: Sequence[Any],
    *,
    exported_by: str,
    exported_at: Optional[datetime] = None,
    brand: Optional[Dict[str, Any]] = None,
    form_settings: Optional[Dict[str, Any]] = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    brand = brand or {}
    dyn_fields = dynamic_field_defs(form_settings)
    headers = list(BASE_HEADERS) + [str(f.get("label") or f.get("id")) for f in dyn_fields]

    navy = _hex(brand.get("header_color") or brand.get("primary_color"), DEFAULT_NAVY)
    gold = _hex(brand.get("secondary_color"), DEFAULT_GOLD)
    alt = _hex(brand.get("table_alt_row_color"), ALT_ROW)
    org_name = brand.get("system_name") or DEFAULT_ORG_NAME
    org_abbr = brand.get("org_abbr") or DEFAULT_ORG_ABBR
    report_title = brand.get("report_title") or DEFAULT_REPORT_TITLE
    sponsor = brand.get("footer_text") or DEFAULT_SPONSOR
    website = brand.get("website") or DEFAULT_WEBSITE
    email = brand.get("email") or DEFAULT_EMAIL
    phone = brand.get("phone") or DEFAULT_PHONE

    company = brand.get("company_name") or ""
    copyright_txt = brand.get("copyright") or ""
    try:
        logo_size = max(40, min(int(brand.get("report_logo_size") or 70), 120))
    except (TypeError, ValueError):
        logo_size = 70

    exported_at = exported_at or datetime.now()
    n_cols = len(headers)
    last_col = get_column_letter(n_cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "الأعضاء"
    ws.sheet_view.rightToLeft = True

    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy_fill = PatternFill("solid", fgColor=navy)
    gold_fill = PatternFill("solid", fgColor=gold)
    header_fill = PatternFill("solid", fgColor=_hex(brand.get("table_header_color"), navy))
    alt_fill = PatternFill("solid", fgColor=alt)
    white_font = Font(name="Arial", bold=True, color=WHITE, size=12)
    title_font = Font(name="Arial", bold=True, color=WHITE, size=18)
    meta_font = Font(name="Arial", bold=True, color=navy, size=10)
    header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

    for r in range(1, 5):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.fill = navy_fill
            cell.alignment = center

    ws.merge_cells(f"A1:{last_col}1")
    ws.merge_cells(f"A2:{last_col}2")
    ws.merge_cells(f"A3:{last_col}3")
    ws.merge_cells(f"A4:{last_col}4")

    ws["A1"] = org_name
    ws["A1"].font = title_font
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = center

    ws["A2"] = org_abbr
    ws["A2"].font = Font(name="Arial", bold=True, color=gold, size=14)
    ws["A2"].fill = navy_fill
    ws["A2"].alignment = center

    ws["A3"] = report_title
    ws["A3"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws["A3"].fill = navy_fill
    ws["A3"].alignment = center

    ws["A4"] = (
        f"المستخدم: {exported_by or '-'}  |  "
        f"التاريخ: {exported_at.strftime('%Y-%m-%d')}  |  "
        f"الوقت: {exported_at.strftime('%H:%M')}  |  "
        f"إجمالي السجلات: {len(items)}"
    )
    ws["A4"].font = Font(name="Arial", color=WHITE, size=10)
    ws["A4"].fill = navy_fill
    ws["A4"].alignment = center

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 22

    logo = _logo_path(brand)
    if logo:
        try:
            img = XLImage(str(logo))
            img.width = logo_size
            img.height = logo_size
            ws.add_image(img, "A1")
        except Exception:
            pass

    strip_row = 5
    ws.merge_cells(f"A{strip_row}:{last_col}{strip_row}")
    ws[f"A{strip_row}"] = sponsor
    ws[f"A{strip_row}"].font = meta_font
    ws[f"A{strip_row}"].fill = gold_fill
    ws[f"A{strip_row}"].alignment = center
    for c in range(1, n_cols + 1):
        ws.cell(row=strip_row, column=c).fill = gold_fill
        ws.cell(row=strip_row, column=c).border = border
    ws.row_dimensions[strip_row].height = 20

    header_row = 6
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 28

    for idx, item in enumerate(items, 1):
        r = header_row + idx
        values = _row_values(idx, item, dyn_fields, form_settings)
        fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = cell_font
            cell.alignment = center if col in (1, 2, 5, 9, 10) else right
            cell.fill = fill
            cell.border = border
        ws.row_dimensions[r].height = 18

    data_end = header_row + max(len(items), 1)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{data_end}"

    min_widths = [5, 12, 18, 14, 13, 10, 10, 14, 11, 11, 16, 16, 14, 16, 18] + [14] * len(dyn_fields)
    for i, base in enumerate(min_widths, 1):
        if i > n_cols:
            break
        letter = get_column_letter(i)
        max_len = base
        for row in ws.iter_rows(min_row=header_row, max_row=data_end, min_col=i, max_col=i):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, min(len(str(val)) + 2, 42))
        ws.column_dimensions[letter].width = max_len

    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.6
    footer_bits = [sponsor, company, website, email, phone, copyright_txt]
    ws.oddFooter.center.text = " | ".join([b for b in footer_bits if b]) + " | صفحة &P من &N"
    ws.evenFooter.center.text = ws.oddFooter.center.text
    ws.oddFooter.center.font = "Arial"
    ws.oddFooter.center.size = 8

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
