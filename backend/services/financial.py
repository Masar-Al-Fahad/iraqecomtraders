"""Financial calculations, audit helpers and report export."""
from __future__ import annotations

import io
import json
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Iterable

from sqlalchemy.ext.asyncio import AsyncSession

from models.financial import FinancialAuditLog

MONEY = Decimal("0.001")
COMMISSION_METHODS = {
    "fixed_per_operation",
    "percentage",
    "monthly_fixed",
    "custom",
}


def calculate_revenue(
    method: str,
    value: Decimal | int | float | str,
    operation_count: int,
    gross_business_value: Decimal | int | float | str = 0,
) -> Decimal:
    """Calculate an immutable monthly revenue snapshot in IQD."""
    if method not in COMMISSION_METHODS:
        raise ValueError("طريقة العمولة غير مدعومة")
    rate = Decimal(str(value or 0))
    operations = max(int(operation_count or 0), 0)
    gross = Decimal(str(gross_business_value or 0))
    if rate < 0 or gross < 0:
        raise ValueError("القيم المالية لا يمكن أن تكون سالبة")
    if method == "fixed_per_operation":
        result = rate * operations
    elif method == "percentage":
        result = gross * rate / Decimal("100")
    elif method == "monthly_fixed":
        result = rate
    else:
        # Custom is deliberately data-driven; the captured value is the amount
        # for this monthly row until a future custom formula is configured.
        result = rate
    return result.quantize(MONEY, rounding=ROUND_HALF_UP)


def json_value(value: Any) -> str | None:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True)


def add_audit(
    db: AsyncSession,
    *,
    action: str,
    entity_type: str,
    actor: str,
    entity_id: int | None = None,
    old_values: Any = None,
    new_values: Any = None,
) -> None:
    db.add(
        FinancialAuditLog(
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            actor=actor,
            old_values=json_value(old_values),
            new_values=json_value(new_values),
        )
    )


def build_financial_xlsx(rows: Iterable[dict], title: str, filters_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    rows = list(rows)
    headers = [
        ("membership_number", "رقم العضوية"),
        ("member_name", "اسم العضو"),
        ("governorate", "المحافظة"),
        ("shipping_operations", "طلبات الشحن"),
        ("shipping_revenue", "إيراد الشحن"),
        ("delivery_operations", "طلبات التوصيل"),
        ("delivery_revenue", "إيراد التوصيل"),
        ("other_operations", "عمليات الخدمات الأخرى"),
        ("other_revenue", "إيراد الخدمات الأخرى"),
        ("total_operations", "إجمالي العمليات"),
        ("total_revenue", "إجمالي إيراد الاتحاد"),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير المالي"
    ws.sheet_view.rightToLeft = True
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2937")
    ws["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = filters_label
    ws["A2"].alignment = Alignment(horizontal="center", readingOrder=2)
    thin = Side(style="thin", color="D1D5DB")
    for col, (_, label) in enumerate(headers, 1):
        cell = ws.cell(4, col, label)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C89B3C")
        cell.alignment = Alignment(horizontal="center", readingOrder=2, wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ridx, row in enumerate(rows, 5):
        for cidx, (key, _) in enumerate(headers, 1):
            value = row.get(key, 0 if "operations" in key or "revenue" in key else "")
            cell = ws.cell(ridx, cidx, float(value) if isinstance(value, Decimal) else value)
            cell.alignment = Alignment(horizontal="center", readingOrder=2)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [15, 24, 14, 14, 16, 14, 16, 18, 18, 16, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{max(4, 4 + len(rows))}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.oddFooter.center.text = "تجمع تجار التجارة الإلكترونية في العراق | صفحة &P من &N"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_erp_xlsx(rows: Iterable[dict], columns: list[tuple[str, str]], title: str, period: str) -> bytes:
    """Build a branded, real XLSX for arbitrary ERP ledgers."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = list(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.sheet_view.rightToLeft = True
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws["A1"] = f"تجمع تجار التجارة الإلكترونية في العراق (MFEC) — {title}"
    ws["A1"].font = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2937")
    ws["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws["A2"] = period
    ws["A2"].alignment = Alignment(horizontal="center", readingOrder=2)
    thin = Side(style="thin", color="D1D5DB")
    for index, (_, label) in enumerate(columns, 1):
        cell = ws.cell(4, index, label)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C89B3C")
        cell.alignment = Alignment(horizontal="center", readingOrder=2)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ridx, row in enumerate(rows, 5):
        for cidx, (key, _) in enumerate(columns, 1):
            value = row.get(key, "")
            cell = ws.cell(ridx, cidx, float(value) if isinstance(value, Decimal) else value)
            cell.alignment = Alignment(horizontal="center", readingOrder=2)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for index, (_, label) in enumerate(columns, 1):
        longest = max([len(str(label)), *[len(str(row.get(columns[index-1][0], ""))) for row in rows]], default=12)
        ws.column_dimensions[get_column_letter(index)].width = min(max(longest + 3, 12), 36)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(columns))}{max(4, 4 + len(rows))}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.oddFooter.center.text = "MFEC | صفحة &P من &N"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
